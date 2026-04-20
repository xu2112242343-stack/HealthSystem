"""从 Excel 导入「仅用于统计」的健康文章：写入 health_articles，show_in_health_guide=False。

**重要**：仅对 Excel 里 **数据库尚不存在** 的 id 执行插入；已存在的文章默认 **跳过**，
避免把原有「健康生活指南」里的条目改成不展示。若确需覆盖已有行，请加 ``--overwrite-existing``。

用户端不展示这些**新插入**的记录；管理端统计与健康内容分布会包含它们。

默认读取工作表 Sheet3（与常见 qianwen_articles.xlsx 一致）；可用 --sheet 指定。

示例::

    cd web/backend
    python -m scripts.import_qianwen_articles --excel \"path/to/qianwen_articles.xlsx\"

可选：--images 指向配图目录；若某行声明了图片但文件缺失则跳过该图并打印警告，不中断导入。

若曾误用旧版脚本把已有文章标成不展示，可用 ``python -m scripts.set_health_guide_visible_range --min-id 101 --max-id 130`` 恢复区间展示。
"""

from __future__ import annotations

import argparse
import mimetypes
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.models import HealthArticle, HealthArticleImage  # noqa: E402
from app.persistence import SessionLocal, init_db  # noqa: E402

REQUIRED_COLUMNS = [
    "id",
    "title",
    "summary",
    "content",
    "disease",
    "type",
    "tags",
    "risk_level",
    "images",
    "image_desc",
    "source",
]


def _split_csv(v: object) -> list[str]:
    if v is None:
        return []
    return [x.strip() for x in str(v).replace("，", ",").split(",") if x.strip()]


def _as_int_id(v: object) -> int:
    if v is None or str(v).strip() == "":
        raise ValueError("id 为空")
    if isinstance(v, bool):
        raise ValueError("id 类型无效")
    if isinstance(v, int):
        return int(v)
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    try:
        return int(float(s))
    except ValueError as e:
        raise ValueError(f"id 无法解析: {v!r}") from e


def _row_to_payload(row_dict: dict[str, object]) -> dict[str, object]:
    article_id = _as_int_id(row_dict["id"])
    payload = {
        "id": article_id,
        "title": str(row_dict.get("title") or "").strip(),
        "summary": str(row_dict.get("summary") or "").strip(),
        "content": str(row_dict.get("content") or "").strip(),
        "disease": str(row_dict.get("disease") or "").strip(),
        "type": str(row_dict.get("type") or "").strip(),
        "tags": str(row_dict.get("tags") or "").strip() or None,
        "risk_level": str(row_dict.get("risk_level") or "").strip() or None,
        "source": str(row_dict.get("source") or "").strip() or None,
        "images": _split_csv(row_dict.get("images")),
        "image_desc": _split_csv(row_dict.get("image_desc")),
    }
    if not payload["title"] or not payload["summary"] or not payload["content"]:
        raise ValueError(f"id={article_id} 缺少 title/summary/content")
    if not payload["disease"] or not payload["type"]:
        raise ValueError(f"id={article_id} 缺少 disease/type")
    return payload


def _resolve_worksheet(wb, sheet: str | None):
    if sheet:
        if sheet not in wb.sheetnames:
            raise RuntimeError(f"没有工作表: {sheet}，现有: {wb.sheetnames}")
        return wb[sheet]

    if "Sheet3" in wb.sheetnames:
        return wb["Sheet3"]

    for name in wb.sheetnames:
        ws = wb[name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        headers = {str(h or "").strip() for h in header_row}
        if "id" in headers and "title" in headers:
            return ws

    raise RuntimeError(
        f"未自动找到数据表（需含 id、title 列）。请用 --sheet 指定，当前工作簿表: {wb.sheetnames}",
    )


def import_excel(
    excel_path: Path,
    images_dir: Path | None,
    *,
    sheet: str | None,
    overwrite_existing: bool,
) -> None:
    init_db()
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = _resolve_worksheet(wb, sheet)

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise RuntimeError("Excel 没有表头")
    headers = [str(h or "").strip() for h in header_row]

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing_cols:
        raise RuntimeError(f"Excel 缺少列: {', '.join(missing_cols)}")

    idx = {name: headers.index(name) for name in headers}
    created = 0
    updated = 0
    skipped_existing = 0
    image_count = 0
    warn_missing_img = 0

    db = SessionLocal()
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all((x is None or str(x).strip() == "") for x in (row or ())):
                continue
            row_dict = {k: row[v] if v < len(row) else None for k, v in idx.items()}
            p = _row_to_payload(row_dict)

            art = db.get(HealthArticle, p["id"])
            is_new = art is None
            if art is None:
                art = HealthArticle(id=p["id"])
                created += 1
            else:
                if not overwrite_existing:
                    skipped_existing += 1
                    continue
                updated += 1

            art.title = p["title"]
            art.summary = p["summary"]
            art.content = p["content"]
            art.disease = p["disease"]
            art.type = p["type"]
            art.tags = p["tags"]
            art.risk_level = p["risk_level"]
            art.source = p["source"]
            art.show_in_health_guide = False

            db.add(art)
            db.flush()

            image_names: list[str] = p["images"]  # type: ignore[assignment]
            image_descs: list[str] = p["image_desc"]  # type: ignore[assignment]

            if images_dir is not None and image_names:
                db.query(HealthArticleImage).filter(HealthArticleImage.article_id == art.id).delete()
                for i, filename in enumerate(image_names, start=1):
                    img_path = images_dir / filename
                    if not img_path.exists():
                        print(f"[WARN] id={art.id} 跳过缺失图片: {img_path}")
                        warn_missing_img += 1
                        continue
                    mime_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
                    desc = image_descs[i - 1] if i - 1 < len(image_descs) else None
                    db.add(
                        HealthArticleImage(
                            article_id=art.id,
                            filename=filename,
                            mime_type=mime_type,
                            image_path=str(img_path.resolve()),
                            image_data=None,
                            image_desc=desc,
                            sort_order=i,
                        ),
                    )
                    image_count += 1
            elif is_new and not image_names:
                pass
            # 未传 --images：保留该文章已有配图（若有）

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    parts = [
        f"导入完成：新增 {created}",
        f"更新 {updated}" if overwrite_existing else "",
        f"跳过已存在 {skipped_existing} 条" if skipped_existing else "",
        "新插入行均为 show_in_health_guide=False",
        f"配图写入 {image_count} 张",
    ]
    if warn_missing_img:
        parts.append(f"缺失跳过 {warn_missing_img} 张")
    print("；".join(p for p in parts if p))


def main() -> None:
    parser = argparse.ArgumentParser(description="导入仅统计用健康文章（新 id 不出现在用户端指南）")
    parser.add_argument("--excel", required=True, help="qianwen_articles.xlsx 路径")
    parser.add_argument(
        "--images",
        default="",
        help="配图目录（可选；不提供则不改写已有文章的图片记录）",
    )
    parser.add_argument("--sheet", default="", help="工作表名（默认识别 Sheet3 或首个含 id/title 的表）")
    parser.add_argument(
        "--overwrite-existing",
        action="store_true",
        help="若 Excel 中的 id 已存在则覆盖内容并设为不在指南展示（慎用，会隐藏原指南文）",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 不存在: {excel_path}")

    images_dir: Path | None = None
    if str(args.images).strip():
        images_dir = Path(args.images).resolve()
        if not images_dir.is_dir():
            raise NotADirectoryError(f"images 不是目录: {images_dir}")

    sheet = str(args.sheet).strip() or None
    import_excel(excel_path, images_dir, sheet=sheet, overwrite_existing=bool(args.overwrite_existing))


if __name__ == "__main__":
    main()
