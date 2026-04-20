from __future__ import annotations

import argparse
import mimetypes
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.models import HealthArticle, HealthArticleImage  # noqa: E402
from app.persistence import SessionLocal, init_db  # noqa: E402


REQUIRED_COLUMNS = [
    "id",
    "title",
    "summary",
    "content",
    "disease",
    "type",
    "tags",
    "risk_level",
    "images",
    "image_desc",
    "source",
]


def _split_csv(v: object) -> list[str]:
    if v is None:
        return []
    return [x.strip() for x in str(v).replace("，", ",").split(",") if x.strip()]


def _row_to_payload(row_dict: dict[str, object]) -> dict[str, object]:
    article_id = int(row_dict["id"])
    payload = {
        "id": article_id,
        "title": str(row_dict.get("title") or "").strip(),
        "summary": str(row_dict.get("summary") or "").strip(),
        "content": str(row_dict.get("content") or "").strip(),
        "disease": str(row_dict.get("disease") or "").strip(),
        "type": str(row_dict.get("type") or "").strip(),
        "tags": str(row_dict.get("tags") or "").strip() or None,
        "risk_level": str(row_dict.get("risk_level") or "").strip() or None,
        "source": str(row_dict.get("source") or "").strip() or None,
        "images": _split_csv(row_dict.get("images")),
        "image_desc": _split_csv(row_dict.get("image_desc")),
    }
    if not payload["title"] or not payload["summary"] or not payload["content"]:
        raise ValueError(f"id={article_id} 缺少 title/summary/content")
    if not payload["disease"] or not payload["type"]:
        raise ValueError(f"id={article_id} 缺少 disease/type")
    return payload


def import_excel(excel_path: Path, images_dir: Path) -> None:
    init_db()
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise RuntimeError("Excel 没有表头")
    headers = [str(h or "").strip() for h in header_row]

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing_cols:
        raise RuntimeError(f"Excel 缺少列: {', '.join(missing_cols)}")

    idx = {name: headers.index(name) for name in headers}
    created = 0
    updated = 0
    image_count = 0

    db = SessionLocal()
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all((x is None or str(x).strip() == "") for x in row):
                continue
            row_dict = {k: row[v] if v < len(row) else None for k, v in idx.items()}
            p = _row_to_payload(row_dict)

            art = db.get(HealthArticle, p["id"])
            if art is None:
                art = HealthArticle(id=p["id"])
                created += 1
            else:
                updated += 1

            art.title = p["title"]
            art.summary = p["summary"]
            art.content = p["content"]
            art.disease = p["disease"]
            art.type = p["type"]
            art.tags = p["tags"]
            art.risk_level = p["risk_level"]
            art.source = p["source"]
            art.show_in_health_guide = True

            db.add(art)
            db.flush()

            db.query(HealthArticleImage).filter(HealthArticleImage.article_id == art.id).delete()

            image_names: list[str] = p["images"]  # type: ignore[assignment]
            image_descs: list[str] = p["image_desc"]  # type: ignore[assignment]
            for i, filename in enumerate(image_names, start=1):
                img_path = images_dir / filename
                if not img_path.exists():
                    raise FileNotFoundError(f"id={art.id} 图片不存在: {img_path}")
                mime_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
                desc = image_descs[i - 1] if i - 1 < len(image_descs) else None
                db.add(
                    HealthArticleImage(
                        article_id=art.id,
                        filename=filename,
                        mime_type=mime_type,
                        image_path=str(img_path.resolve()),
                        image_data=None,
                        image_desc=desc,
                        sort_order=i,
                    ),
                )
                image_count += 1

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    print(f"导入完成：新增文章 {created}，更新文章 {updated}，图片 {image_count}")


def main() -> None:
    parser = argparse.ArgumentParser(description="导入健康生活指南 Excel+图片到数据库（数据库仅存图片路径）")
    parser.add_argument("--excel", required=True, help="articles.xlsx 路径")
    parser.add_argument("--images", required=True, help="images 文件夹路径")
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    images_dir = Path(args.images).resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 不存在: {excel_path}")
    if not images_dir.exists():
        raise FileNotFoundError(f"images 目录不存在: {images_dir}")

    import_excel(excel_path, images_dir)


if __name__ == "__main__":
    main()

